"""Build the Fase 4 multi-source surfaceome universe."""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import gzip
import hashlib
import io
import os
import re
import sys
import urllib.request
import zipfile
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import openpyxl
import yaml


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from src.utils.matplotlib_repro import configure_reproducible_svg, save_svg

RAW_DIR = REPO_ROOT / "data" / "raw"
PROCESSED_DIR = REPO_ROOT / "data" / "processed"
CHECKSUM_DIR = REPO_ROOT / "data" / "checksums"
RESULTS_DIR = REPO_ROOT / "results" / "tables"
FIGURES_DIR = REPO_ROOT / "results" / "figures"
DOCS_DIR = REPO_ROOT / "docs"
PROVENANCE_LOG = DOCS_DIR / "provenance_log.tsv"
configure_reproducible_svg()

ID_MAP = PROCESSED_DIR / "id_map_master.tsv"
CONTROLS_YAML = REPO_ROOT / "config" / "controls.yaml"
UNIPROT_TOPOLOGY = RAW_DIR / "uniprot" / "uniprot_reviewed_human_topology.tsv.gz"
UNIPROT_GPI = RAW_DIR / "uniprot" / "uniprot_reviewed_human_gpi.tsv.gz"
HPA_SUBCELLULAR = RAW_DIR / "hpa" / "subcellular_location.tsv.zip"

USER_AGENT = "surfaceome-gastric-cancer-phase4/0.1"

GO_SURFACE_IDS = {"GO:0009986", "GO:0005886", "GO:0005887"}
GO_SURFACE_TERMS = {
    "cell surface",
    "plasma membrane",
    "integral component of plasma membrane",
}

PHASE4_DOWNLOADS = [
    {
        "source_id": "cancer_surfaceome_atlas_tcsa",
        "url": "https://static-content.springer.com/esm/art%3A10.1038%2Fs43018-021-00282-w/MediaObjects/43018_2021_282_MOESM2_ESM.xlsx",
        "local_path": RAW_DIR / "surfaceome" / "tcsa_supplementary_tables_1_40.xlsx",
        "version_or_release": "Hu et al. Nature Cancer 2021 Supplementary Tables 1-40; Last-Modified Thu, 26 Oct 2023 09:19:51 GMT",
        "license_or_terms": "Nature/Springer supplementary material terms",
        "notes": "Cancer Surfaceome Atlas supplementary workbook; Table S2 contains 3,567 GESPs.",
    },
    {
        "source_id": "cspa",
        "url": "https://journals.plos.org/plosone/article/file?type=supplementary&id=10.1371/journal.pone.0121314.s003",
        "local_path": RAW_DIR / "surfaceome" / "cspa_pone_0121314_s003.xlsx",
        "version_or_release": "Bausch-Fluck et al. PLOS ONE 2015 S2 file; Last-Modified Mon, 30 Nov 2020 21:58:07 GMT",
        "license_or_terms": "PLOS ONE article supplementary material terms",
        "notes": "CSPA S2 file; Table A contains human experimentally observed surfaceome proteins.",
    },
    {
        "source_id": "surfy",
        "url": "https://wlab.ethz.ch/surfaceome/table_S3_surfaceome.xlsx",
        "local_path": RAW_DIR / "surfaceome" / "surfy_table_s3_surfaceome.xlsx",
        "version_or_release": "SURFY table S3; Last-Modified Mon, 09 Mar 2026 13:42:02 GMT",
        "license_or_terms": "Wollscheid lab SURFY download terms; verify before redistribution",
        "notes": "In silico human surfaceome workbook; sheet 'in silico surfaceome only' contains predicted surface proteins.",
    },
    {
        "source_id": "uniprot_reviewed_human_go",
        "url": "https://rest.uniprot.org/uniprotkb/stream?compressed=true&format=tsv&fields=accession,gene_names,go_id,go_c&query=(reviewed:true)%20AND%20(organism_id:9606)",
        "local_path": RAW_DIR / "uniprot" / "uniprot_reviewed_human_go.tsv.gz",
        "version_or_release": "UniProt reviewed human GO cellular component stream; release observed as 2026_01 in Fase 1",
        "license_or_terms": "UniProt terms",
        "notes": "Reviewed human UniProt accessions with GO cellular component terms and GO IDs.",
    },
    {
        "source_id": "uniprot_reviewed_human_gpi_lipidation",
        "url": "https://rest.uniprot.org/uniprotkb/stream?compressed=true&format=tsv&fields=accession,gene_names,ft_lipid,cc_subcellular_location&query=(reviewed:true)%20AND%20(organism_id:9606)",
        "local_path": UNIPROT_GPI,
        "version_or_release": "UniProt reviewed human lipidation/subcellular stream; release recorded by checksum manifest at retrieval",
        "license_or_terms": "UniProt terms",
        "notes": "Reviewed human UniProt accessions with lipidation and subcellular fields for direct GPI-anchor evidence.",
    },
]

SOURCE_CHECKSUM_MANIFESTS = [
    CHECKSUM_DIR / "xena_toil_sha256.tsv",
    CHECKSUM_DIR / "hpa_sha256.tsv",
    CHECKSUM_DIR / "uniprot_sha256.tsv",
    CHECKSUM_DIR / "gdc_tcga_stad_sha256.tsv",
    CHECKSUM_DIR / "hgnc_sha256.tsv",
    CHECKSUM_DIR / "surfaceome_sources_sha256.tsv",
]


@dataclass
class IdMaps:
    rows: list[dict[str, str]]
    by_symbol: dict[str, dict[str, str]]
    by_alias: dict[str, str]
    by_ensembl: dict[str, str]
    by_uniprot: dict[str, str]
    by_entrez: dict[str, str]


def request(url: str, method: str = "GET", timeout: int = 180) -> urllib.request.addinfourl:
    headers = {"User-Agent": USER_AGENT, "Accept": "*/*"}
    return urllib.request.urlopen(urllib.request.Request(url, headers=headers, method=method), timeout=timeout)


def fetch_head(url: str, timeout: int = 60) -> dict[str, str]:
    try:
        with request(url, method="HEAD", timeout=timeout) as response:
            return {key.lower(): value for key, value in response.headers.items()}
    except Exception:
        return {}


def sha256_file(path: Path, chunk_size: int = 1024 * 1024 * 8) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(chunk_size), b""):
            digest.update(chunk)
    return digest.hexdigest()


def relative(path: Path) -> str:
    return path.resolve().relative_to(REPO_ROOT.resolve()).as_posix()


def write_tsv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, delimiter="\t", fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def download_sources(force: bool, timeout: int) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    for spec in PHASE4_DOWNLOADS:
        path = Path(spec["local_path"])
        path.parent.mkdir(parents=True, exist_ok=True)
        headers = fetch_head(str(spec["url"]), timeout=timeout)
        action = "verified_existing_raw"
        if force or not path.exists():
            action = "downloaded_raw"
            print(f"Downloading {spec['source_id']}: {path.name}", flush=True)
            part = path.with_name(path.name + ".part")
            with request(str(spec["url"]), timeout=timeout) as response, part.open("wb") as handle:
                if not headers:
                    headers = {key.lower(): value for key, value in response.headers.items()}
                for chunk in iter(lambda: response.read(1024 * 1024 * 4), b""):
                    handle.write(chunk)
            os.replace(part, path)
        records.append(
            {
                "source_id": spec["source_id"],
                "action": action,
                "local_path": relative(path),
                "filename": path.name,
                "url_or_endpoint": spec["url"],
                "retrieval_date": dt.date.today().isoformat(),
                "version_or_release": spec["version_or_release"],
                "bytes": path.stat().st_size,
                "sha256": sha256_file(path),
                "status": "ok",
                "last_modified": headers.get("last-modified", ""),
                "content_length": headers.get("content-length", ""),
                "license_or_terms": spec["license_or_terms"],
                "notes": spec["notes"],
            }
        )
    return records


def update_checksum_manifests(records: list[dict[str, object]]) -> None:
    fields = [
        "source_id",
        "action",
        "local_path",
        "filename",
        "url_or_endpoint",
        "retrieval_date",
        "version_or_release",
        "bytes",
        "sha256",
        "status",
        "last_modified",
        "content_length",
        "license_or_terms",
        "notes",
    ]
    write_tsv(CHECKSUM_DIR / "surfaceome_sources_sha256.tsv", records, fields)
    checksums: dict[str, str] = {}
    checksum_path = CHECKSUM_DIR / "sha256sums.txt"
    if checksum_path.exists():
        with checksum_path.open("r", encoding="utf-8") as handle:
            for line in handle:
                stripped = line.strip()
                if stripped:
                    checksum, rel_path = stripped.split(maxsplit=1)
                    checksums[rel_path.strip()] = checksum
    for manifest in SOURCE_CHECKSUM_MANIFESTS:
        if not manifest.exists():
            continue
        for row in read_tsv(manifest):
            rel_path = row.get("local_path", "")
            checksum = row.get("sha256", "")
            if rel_path and checksum:
                checksums[rel_path] = checksum
    for record in records:
        checksums[str(record["local_path"])] = str(record["sha256"])
    with checksum_path.open("w", encoding="utf-8", newline="") as handle:
        for rel_path in sorted(checksums):
            handle.write(f"{checksums[rel_path]}  {rel_path}\n")


def append_provenance(records: list[dict[str, object]]) -> None:
    fieldnames = [
        "date",
        "source_id",
        "action",
        "file_or_endpoint",
        "version_or_release",
        "checksum_or_hash",
        "notes",
    ]
    existing: set[tuple[str, str, str]] = set()
    if PROVENANCE_LOG.exists():
        with PROVENANCE_LOG.open("r", encoding="utf-8", newline="") as handle:
            for row in csv.DictReader(handle, delimiter="\t"):
                existing.add((row.get("source_id", ""), row.get("file_or_endpoint", ""), row.get("checksum_or_hash", "")))
    write_header = not PROVENANCE_LOG.exists() or PROVENANCE_LOG.stat().st_size == 0
    with PROVENANCE_LOG.open("a", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, delimiter="\t", fieldnames=fieldnames, lineterminator="\n")
        if write_header:
            writer.writeheader()
        for record in records:
            key = (str(record["source_id"]), str(record["url_or_endpoint"]), str(record["sha256"]))
            if key in existing:
                continue
            writer.writerow(
                {
                    "date": record["retrieval_date"],
                    "source_id": record["source_id"],
                    "action": record["action"],
                    "file_or_endpoint": record["url_or_endpoint"],
                    "version_or_release": record["version_or_release"],
                    "checksum_or_hash": record["sha256"],
                    "notes": f"{record['local_path']}; {record['notes']}",
                }
            )


def read_tsv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle, delimiter="\t"))


def load_id_maps() -> IdMaps:
    rows = read_tsv(ID_MAP)
    by_symbol: dict[str, dict[str, str]] = {}
    by_alias: dict[str, str] = {}
    by_ensembl: dict[str, str] = {}
    by_uniprot: dict[str, str] = {}
    by_entrez: dict[str, str] = {}
    for row in rows:
        symbol = row["hgnc_symbol"]
        by_symbol[symbol] = row
        for alias_field in ["source_gene_symbol", "alias_symbols", "previous_symbols"]:
            for alias in str(row.get(alias_field, "")).replace(";", "|").split("|"):
                clean_alias = alias.strip()
                if clean_alias:
                    by_alias.setdefault(clean_alias, symbol)
        if row.get("ensembl_gene_id"):
            by_ensembl[row["ensembl_gene_id"]] = symbol
        primary_accession = row.get("uniprot_accession", "")
        if primary_accession:
            by_uniprot[primary_accession] = symbol
        for accession in str(row.get("alternative_uniprot_accessions", "")).split("|"):
            if accession:
                by_uniprot.setdefault(accession, symbol)
        if row.get("entrez_id"):
            by_entrez[str(row["entrez_id"])] = symbol
    return IdMaps(rows=rows, by_symbol=by_symbol, by_alias=by_alias, by_ensembl=by_ensembl, by_uniprot=by_uniprot, by_entrez=by_entrez)


def resolve_symbol(id_maps: IdMaps, symbol: object = None, ensembl: object = None, uniprot: object = None, entrez: object = None) -> str:
    if symbol:
        clean = str(symbol).strip()
        if clean in id_maps.by_symbol:
            return clean
        if clean in id_maps.by_alias:
            return id_maps.by_alias[clean]
    if ensembl:
        clean = str(ensembl).strip().split(".", 1)[0]
        if clean in id_maps.by_ensembl:
            return id_maps.by_ensembl[clean]
    if uniprot:
        clean = str(uniprot).strip()
        if clean in id_maps.by_uniprot:
            return id_maps.by_uniprot[clean]
    if entrez:
        clean = str(entrez).strip()
        if clean.endswith(".0"):
            clean = clean[:-2]
        if clean in id_maps.by_entrez:
            return id_maps.by_entrez[clean]
    return ""


def resolve_symbol_from_gene_names(id_maps: IdMaps, gene_names: object) -> str:
    for token in str(gene_names or "").replace(";", " ").split():
        symbol = resolve_symbol(id_maps, symbol=token)
        if symbol:
            return symbol
    return ""


def workbook(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def parse_tcsa(id_maps: IdMaps) -> tuple[set[str], dict[str, dict[str, object]]]:
    path = RAW_DIR / "surfaceome" / "tcsa_supplementary_tables_1_40.xlsx"
    wb = workbook(path)
    ws = wb["Table S2"]
    genes: set[str] = set()
    meta: dict[str, dict[str, object]] = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or not row[0]:
            continue
        symbol = resolve_symbol(id_maps, symbol=row[1], ensembl=row[0])
        if not symbol:
            continue
        genes.add(symbol)
        meta[symbol] = {
            "tcsa_ensembl_gene_id": str(row[0]),
            "tcsa_core_gesp_score": row[3] if len(row) > 3 else "",
            "tcsa_final_gesp_score": row[4] if len(row) > 4 else "",
            "tcsa_integral_monotopic": row[5] if len(row) > 5 else "",
        }
    return genes, meta


def parse_cspa(id_maps: IdMaps) -> tuple[set[str], set[str], dict[str, dict[str, object]]]:
    path = RAW_DIR / "surfaceome" / "cspa_pone_0121314_s003.xlsx"
    wb = workbook(path)
    ws = wb["Table A"]
    positive: set[str] = set()
    unspecific: set[str] = set()
    meta: dict[str, dict[str, object]] = {}
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        symbol = resolve_symbol(
            id_maps,
            symbol=row[idx.get("ENTREZ gene symbol")],
            uniprot=row[idx.get("ID_link")],
            entrez=row[idx.get("ENTREZ_gene_ID")],
        )
        if not symbol:
            continue
        category = str(row[idx["CSPA category"]] or "")
        if category.startswith("1") or category.startswith("2"):
            positive.add(symbol)
        else:
            unspecific.add(symbol)
        meta[symbol] = {
            "cspa_category": category,
            "cspa_detection_cell_types": row[idx.get("count detection in different cell types")],
            "cspa_uniprot": row[idx.get("ID_link")],
        }
    return positive, unspecific, meta


def parse_surfy(id_maps: IdMaps) -> tuple[set[str], dict[str, dict[str, object]]]:
    path = RAW_DIR / "surfaceome" / "surfy_table_s3_surfaceome.xlsx"
    wb = workbook(path)
    ws = wb["in silico surfaceome only"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    idx = {name: i for i, name in enumerate(headers)}
    genes: set[str] = set()
    meta: dict[str, dict[str, object]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        symbol = resolve_symbol(id_maps, symbol=row[idx["UniProt gene"]], uniprot=row[idx["UniProt accession"]])
        if not symbol:
            continue
        genes.add(symbol)
        meta[symbol] = {
            "surfy_label_source": row[idx.get("Surfaceome Label Source")],
            "surfy_tm_domains": row[idx.get("TM domains")],
            "surfy_topology": row[idx.get("topology")],
        }
    return genes, meta


def parse_uniprot_topology(id_maps: IdMaps) -> dict[str, dict[str, bool]]:
    by_symbol: dict[str, dict[str, bool]] = defaultdict(lambda: defaultdict(bool))
    with gzip.open(UNIPROT_TOPOLOGY, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            symbol = resolve_symbol(id_maps, uniprot=row.get("Entry"), symbol=(row.get("Gene Names", "").split() or [""])[0])
            if not symbol:
                continue
            topology = str(row.get("Topological domain", ""))
            transmembrane = str(row.get("Transmembrane", ""))
            signal = str(row.get("Signal peptide", ""))
            by_symbol[symbol]["uniprot_extracellular_topology"] = by_symbol[symbol]["uniprot_extracellular_topology"] or (
                "extracellular" in topology.lower()
            )
            by_symbol[symbol]["uniprot_transmembrane"] = by_symbol[symbol]["uniprot_transmembrane"] or bool(transmembrane)
            by_symbol[symbol]["uniprot_signal_peptide"] = by_symbol[symbol]["uniprot_signal_peptide"] or bool(signal)
    return {key: dict(value) for key, value in by_symbol.items()}


def parse_uniprot_gpi(id_maps: IdMaps) -> tuple[set[str], set[str], dict[str, dict[str, str]]]:
    direct_lipid: set[str] = set()
    subcellular_only: set[str] = set()
    meta: dict[str, dict[str, str]] = {}
    with gzip.open(UNIPROT_GPI, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            symbol = resolve_symbol(id_maps, uniprot=row.get("Entry")) or resolve_symbol_from_gene_names(id_maps, row.get("Gene Names", ""))
            if not symbol:
                continue
            lipidation = str(row.get("Lipidation", ""))
            subcellular = str(row.get("Subcellular location [CC]", ""))
            has_direct_gpi = "gpi-anchor" in lipidation.lower() or "gpi anchor" in lipidation.lower()
            has_subcellular_gpi = "gpi-anchor" in subcellular.lower() or "gpi anchor" in subcellular.lower()
            if not has_direct_gpi and not has_subcellular_gpi:
                continue
            if has_direct_gpi:
                direct_lipid.add(symbol)
                evidence_class = "confirmed_uniprot_lipid_gpi_anchor"
            elif symbol not in direct_lipid:
                subcellular_only.add(symbol)
                evidence_class = "subcellular_gpi_without_lipid_feature"
            else:
                continue
            meta[symbol] = {
                "uniprot_gpi_accession": row.get("Entry", ""),
                "uniprot_gpi_evidence_class": evidence_class,
                "uniprot_gpi_lipidation": lipidation,
                "uniprot_gpi_subcellular_location": subcellular,
            }
    return direct_lipid, subcellular_only - direct_lipid, meta


def parse_uniprot_go(id_maps: IdMaps) -> tuple[set[str], dict[str, str]]:
    path = RAW_DIR / "uniprot" / "uniprot_reviewed_human_go.tsv.gz"
    genes: set[str] = set()
    terms_by_symbol: dict[str, set[str]] = defaultdict(set)
    with gzip.open(path, "rt", encoding="utf-8", errors="replace", newline="") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            symbol = resolve_symbol(id_maps, uniprot=row.get("Entry"), symbol=(row.get("Gene Names", "").split() or [""])[0])
            if not symbol:
                continue
            go_ids = set(re.findall(r"GO:\d+", str(row.get("Gene Ontology IDs", ""))))
            go_terms = {term.strip().lower() for term in str(row.get("Gene Ontology (cellular component)", "")).split(";")}
            if go_ids & GO_SURFACE_IDS or GO_SURFACE_TERMS & go_terms:
                genes.add(symbol)
                terms_by_symbol[symbol].update(sorted((go_ids & GO_SURFACE_IDS) | (GO_SURFACE_TERMS & go_terms)))
    return genes, {symbol: "|".join(sorted(terms)) for symbol, terms in terms_by_symbol.items()}


def parse_hpa_subcellular(id_maps: IdMaps) -> tuple[set[str], dict[str, dict[str, object]], set[str], set[str]]:
    membrane_genes: set[str] = set()
    intracellular_exclusive: set[str] = set()
    extracellular_genes: set[str] = set()
    meta: dict[str, dict[str, object]] = {}
    surface_terms = {"plasma membrane", "cell membrane"}
    intracellular_terms = {
        "nucleoplasm",
        "nucleus",
        "nucleoli",
        "nucleoli fibrillar center",
        "mitochondria",
        "cytosol",
        "centrosome",
        "microtubules",
        "actin filaments",
        "intermediate filaments",
        "endoplasmic reticulum",
        "golgi apparatus",
        "vesicles",
        "peroxisomes",
    }
    with zipfile.ZipFile(HPA_SUBCELLULAR) as archive:
        with archive.open(archive.namelist()[0]) as handle:
            text = io.TextIOWrapper(handle, encoding="utf-8", errors="replace")
            for row in csv.DictReader(text, delimiter="\t"):
                symbol = resolve_symbol(id_maps, ensembl=row.get("Gene"), symbol=row.get("Gene name"))
                if not symbol:
                    continue
                fields = [
                    row.get("Main location", ""),
                    row.get("Additional location", ""),
                    row.get("Approved", ""),
                    row.get("Supported", ""),
                    row.get("Enhanced", ""),
                ]
                locations = {
                    item.strip().lower()
                    for field in fields
                    for item in str(field or "").replace(";", "|").split("|")
                    if item.strip()
                }
                extracellular = str(row.get("Extracellular location", "") or "").strip()
                if any(any(term in loc for term in surface_terms) for loc in locations):
                    membrane_genes.add(symbol)
                if extracellular:
                    extracellular_genes.add(symbol)
                if locations and all(any(term == loc or term in loc for term in intracellular_terms) for loc in locations):
                    intracellular_exclusive.add(symbol)
                meta[symbol] = {
                    "hpa_main_location": row.get("Main location", ""),
                    "hpa_additional_location": row.get("Additional location", ""),
                    "hpa_reliability": row.get("Reliability", ""),
                    "hpa_extracellular_location": extracellular,
                }
    return membrane_genes, meta, intracellular_exclusive, extracellular_genes


def load_control_symbols() -> tuple[set[str], set[str], set[str]]:
    controls = yaml.safe_load(CONTROLS_YAML.read_text(encoding="utf-8"))
    positive = {item["gene"] for item in controls.get("positive_controls", [])}
    positive.update(item["gene"] for item in controls.get("secondary_benchmark_targets", []))
    negative = {item["gene"] for item in controls.get("negative_controls_intracellular_or_secreted", [])}
    tme = {item["gene"] for item in controls.get("tme_or_off_tumor_penalty_controls", [])}
    return positive, negative, tme


def category_for(
    score: int,
    support_count: int,
    has_strong: bool,
    has_anchor: bool,
    secreted_only: bool,
    intracellular_only: bool,
    has_any_support: bool,
    has_tm_or_signal: bool,
) -> str:
    if score >= 6 and support_count >= 3 and has_strong and has_anchor and not secreted_only and not intracellular_only:
        return "core_surfaceome"
    if score >= 5 and support_count >= 4 and has_strong and has_anchor and not secreted_only and not intracellular_only:
        return "probable_surfaceome"
    if has_any_support or has_tm_or_signal:
        return "ambiguous_membrane_or_surface_context"
    return "excluded"


def build_universe() -> tuple[list[dict[str, object]], dict[str, set[str]], list[dict[str, object]], list[dict[str, object]]]:
    id_maps = load_id_maps()
    tcsa, tcsa_meta = parse_tcsa(id_maps)
    cspa, cspa_unspecific, cspa_meta = parse_cspa(id_maps)
    surfy, surfy_meta = parse_surfy(id_maps)
    uniprot_topology = parse_uniprot_topology(id_maps)
    uniprot_gpi, uniprot_gpi_subcellular_only, uniprot_gpi_meta = parse_uniprot_gpi(id_maps)
    go, go_terms = parse_uniprot_go(id_maps)
    hpa_membrane, hpa_meta, hpa_intracellular, hpa_extracellular = parse_hpa_subcellular(id_maps)
    positive_controls, negative_controls, tme_controls = load_control_symbols()

    rows: list[dict[str, object]] = []
    for id_row in id_maps.rows:
        symbol = id_row["hgnc_symbol"]
        topo = uniprot_topology.get(symbol, {})
        in_tcsa = symbol in tcsa
        in_cspa = symbol in cspa
        in_surfy = symbol in surfy
        has_uniprot_ext = bool(topo.get("uniprot_extracellular_topology"))
        has_uniprot_tm = bool(topo.get("uniprot_transmembrane"))
        has_signal = bool(topo.get("uniprot_signal_peptide"))
        has_gpi_direct = symbol in uniprot_gpi
        has_gpi_subcellular_only = symbol in uniprot_gpi_subcellular_only
        has_go = symbol in go
        has_hpa = symbol in hpa_membrane
        has_secreted_evidence = symbol in hpa_extracellular or (has_signal and not has_uniprot_tm and not has_uniprot_ext)
        support_count = sum([in_tcsa, in_cspa, in_surfy, has_uniprot_ext, has_gpi_direct, has_go, has_hpa])
        curated_or_experimental = in_tcsa or in_cspa
        has_anchor = has_uniprot_ext or has_uniprot_tm or has_gpi_direct or has_hpa or in_surfy
        score = 0
        if curated_or_experimental:
            score += 3
        if has_uniprot_ext:
            score += 2
        if has_gpi_direct:
            score += 2
        if has_go:
            score += 1
        if has_hpa:
            score += 1
        if in_surfy:
            score += 1

        intracellular_only = symbol in hpa_intracellular and support_count == 0
        secreted_only = has_secreted_evidence and not has_uniprot_tm and not has_uniprot_ext and not has_gpi_direct and not (
            in_tcsa or in_cspa or in_surfy or has_hpa or has_go
        )
        if intracellular_only:
            score -= 2
        if secreted_only:
            score -= 2

        category = category_for(
            score=score,
            support_count=support_count,
            has_strong=curated_or_experimental or has_uniprot_ext or has_gpi_direct,
            has_anchor=has_anchor,
            secreted_only=secreted_only,
            intracellular_only=intracellular_only,
            has_any_support=support_count > 0,
            has_tm_or_signal=has_uniprot_tm or has_signal,
        )
        flags = []
        if secreted_only:
            flags.append("secreted_only_without_surface_anchor")
        if intracellular_only:
            flags.append("exclusive_intracellular_hpa_localization")
        if (in_tcsa or in_cspa or has_go) and not has_anchor:
            flags.append("surface_annotation_without_membrane_anchor_or_hpa_membrane")
        if has_gpi_subcellular_only:
            flags.append("uniprot_gpi_subcellular_only_not_counted_as_direct_lipid_evidence")
        if symbol in cspa_unspecific and symbol not in cspa:
            flags.append("cspa_unspecific_copurified_not_counted_as_surface_support")
        if symbol in tme_controls:
            flags.append("tme_or_off_tumor_penalty_control")
        if symbol in negative_controls and category in {"core_surfaceome", "probable_surfaceome"}:
            flags.append("negative_control_in_core_probable_requires_review")
        if symbol in positive_controls and category == "excluded":
            flags.append("positive_control_excluded_requires_review")

        source_list = [
            name
            for name, present in [
                ("tcsa", in_tcsa),
                ("cspa", in_cspa),
                ("surfy", in_surfy),
                ("uniprot_extracellular_topology", has_uniprot_ext),
                ("uniprot_gpi_anchor", has_gpi_direct),
                ("go_surface_or_plasma_membrane", has_go),
                ("hpa_plasma_membrane", has_hpa),
            ]
            if present
        ]
        row = {
            "hgnc_symbol": symbol,
            "ensembl_gene_id": id_row.get("ensembl_gene_id", ""),
            "uniprot_accession": id_row.get("uniprot_accession", ""),
            "protein_name": id_row.get("protein_name", ""),
            "surfaceome_confidence_score": score,
            "surfaceome_category": category,
            "surface_support_source_count": support_count,
            "surface_support_sources": "|".join(source_list),
            "in_tcsa": str(in_tcsa).lower(),
            "in_cspa": str(in_cspa).lower(),
            "in_surfy": str(in_surfy).lower(),
            "uniprot_extracellular_topology": str(has_uniprot_ext).lower(),
            "uniprot_gpi_anchor": str(has_gpi_direct).lower(),
            "uniprot_gpi_evidence_class": uniprot_gpi_meta.get(symbol, {}).get("uniprot_gpi_evidence_class", ""),
            "uniprot_gpi_subcellular_only": str(has_gpi_subcellular_only).lower(),
            "uniprot_transmembrane": str(has_uniprot_tm).lower(),
            "uniprot_signal_peptide": str(has_signal).lower(),
            "go_surface_or_plasma_membrane": str(has_go).lower(),
            "go_surface_terms": go_terms.get(symbol, ""),
            "hpa_plasma_membrane": str(has_hpa).lower(),
            "hpa_intracellular_only_flag": str(intracellular_only).lower(),
            "secreted_only_flag": str(secreted_only).lower(),
            "control_role": (
                "positive_or_benchmark"
                if symbol in positive_controls
                else "negative_intracellular_or_secreted"
                if symbol in negative_controls
                else "tme_or_off_tumor"
                if symbol in tme_controls
                else ""
            ),
            "interpretation_flags": "|".join(flags),
            "tcsa_core_gesp_score": tcsa_meta.get(symbol, {}).get("tcsa_core_gesp_score", ""),
            "tcsa_final_gesp_score": tcsa_meta.get(symbol, {}).get("tcsa_final_gesp_score", ""),
            "cspa_category": cspa_meta.get(symbol, {}).get("cspa_category", ""),
            "cspa_detection_cell_types": cspa_meta.get(symbol, {}).get("cspa_detection_cell_types", ""),
            "surfy_label_source": surfy_meta.get(symbol, {}).get("surfy_label_source", ""),
            "surfy_tm_domains": surfy_meta.get(symbol, {}).get("surfy_tm_domains", ""),
            "hpa_main_location": hpa_meta.get(symbol, {}).get("hpa_main_location", ""),
            "hpa_additional_location": hpa_meta.get(symbol, {}).get("hpa_additional_location", ""),
            "hpa_reliability": hpa_meta.get(symbol, {}).get("hpa_reliability", ""),
            "uniprot_gpi_accession": uniprot_gpi_meta.get(symbol, {}).get("uniprot_gpi_accession", ""),
            "uniprot_gpi_lipidation": uniprot_gpi_meta.get(symbol, {}).get("uniprot_gpi_lipidation", ""),
        }
        rows.append(row)

    sets = {
        "tcsa": tcsa,
        "cspa": cspa,
        "surfy": surfy,
        "uniprot_extracellular_topology": {s for s, topo in uniprot_topology.items() if topo.get("uniprot_extracellular_topology")},
        "uniprot_gpi_anchor": uniprot_gpi,
        "go_surface_or_plasma_membrane": go,
        "hpa_plasma_membrane": hpa_membrane,
    }
    summary_rows, jaccard_rows = build_summary_rows(rows, sets)
    return rows, sets, summary_rows, jaccard_rows


def build_summary_rows(rows: list[dict[str, object]], sets: dict[str, set[str]]) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    own = {str(row["hgnc_symbol"]) for row in rows if row["surfaceome_category"] in {"core_surfaceome", "probable_surfaceome"}}
    category_counts = Counter(str(row["surfaceome_category"]) for row in rows)
    summary_rows = [
        {"summary_type": "category", "label": category, "n_genes": count, "notes": ""}
        for category, count in sorted(category_counts.items())
    ]
    for source, genes in sorted(sets.items()):
        summary_rows.append(
            {
                "summary_type": "source",
                "label": source,
                "n_genes": len(genes),
                "notes": f"{len(genes & own)} overlap with core+probable universe",
            }
        )

    published = {key: sets[key] for key in ["tcsa", "cspa", "surfy"]}
    jaccard_rows = []
    for source, genes in published.items():
        union = own | genes
        intersection = own & genes
        jaccard_rows.append(
            {
                "published_list": source,
                "own_core_probable_n": len(own),
                "published_n": len(genes),
                "intersection_n": len(intersection),
                "union_n": len(union),
                "jaccard": (len(intersection) / len(union)) if union else 0.0,
            }
        )
    return summary_rows, jaccard_rows


def build_audit_rows(rows: list[dict[str, object]], sets: dict[str, set[str]]) -> list[dict[str, object]]:
    positive_controls, negative_controls, tme_controls = load_control_symbols()
    row_by_symbol = {str(row["hgnc_symbol"]): row for row in rows}
    own = {str(row["hgnc_symbol"]) for row in rows if row["surfaceome_category"] in {"core_surfaceome", "probable_surfaceome"}}
    published_union = sets["tcsa"] | sets["cspa"] | sets["surfy"]
    audit: list[dict[str, object]] = []
    for symbol in sorted(positive_controls | negative_controls | tme_controls):
        row = row_by_symbol.get(symbol, {})
        audit.append(
            {
                "audit_type": "control",
                "hgnc_symbol": symbol,
                "surfaceome_category": row.get("surfaceome_category", "missing"),
                "surfaceome_confidence_score": row.get("surfaceome_confidence_score", ""),
                "control_role": row.get("control_role", ""),
                "status": "ok"
                if (
                    (symbol in positive_controls and row.get("surfaceome_category") != "excluded")
                    or (symbol in negative_controls and row.get("surfaceome_category") not in {"core_surfaceome", "probable_surfaceome"})
                    or symbol in tme_controls
                )
                else "review_required",
                "notes": row.get("interpretation_flags", ""),
            }
        )
    for symbol in sorted(own - published_union):
        row = row_by_symbol[symbol]
        audit.append(
            {
                "audit_type": "potential_false_positive_not_in_published_lists",
                "hgnc_symbol": symbol,
                "surfaceome_category": row["surfaceome_category"],
                "surfaceome_confidence_score": row["surfaceome_confidence_score"],
                "control_role": row["control_role"],
                "status": "review_if_top_candidate",
                "notes": row["surface_support_sources"],
            }
        )
    for symbol in sorted(published_union - own):
        row = row_by_symbol.get(symbol, {})
        audit.append(
            {
                "audit_type": "potential_false_negative_published_absent_from_core_probable",
                "hgnc_symbol": symbol,
                "surfaceome_category": row.get("surfaceome_category", "not_in_id_map"),
                "surfaceome_confidence_score": row.get("surfaceome_confidence_score", ""),
                "control_role": row.get("control_role", ""),
                "status": "review_if_biologically_relevant",
                "notes": row.get("interpretation_flags", ""),
            }
        )
    return audit


def plot_source_overlap(summary_rows: list[dict[str, object]], output: Path) -> None:
    source_rows = [row for row in summary_rows if row["summary_type"] == "source"]
    labels = [str(row["label"]) for row in source_rows]
    totals = [int(row["n_genes"]) for row in source_rows]
    overlaps = [int(re.match(r"(\d+)", str(row["notes"])).group(1)) for row in source_rows]
    fig, ax = plt.subplots(figsize=(8.5, 4.8))
    x = range(len(labels))
    ax.bar(x, totals, color="#94a3b8", label="source list")
    ax.bar(x, overlaps, color="#2563eb", label="overlap with core+probable")
    ax.set_xticks(list(x))
    ax.set_xticklabels(labels, rotation=30, ha="right")
    ax.set_ylabel("Genes")
    ax.set_title("Surfaceome source coverage and overlap")
    ax.legend(frameon=False)
    ax.grid(axis="y", linewidth=0.35, alpha=0.3)
    fig.tight_layout()
    output.parent.mkdir(parents=True, exist_ok=True)
    save_svg(fig, output)
    plt.close(fig)


def plot_jaccard(jaccard_rows: list[dict[str, object]], output: Path) -> None:
    labels = [str(row["published_list"]) for row in jaccard_rows]
    values = [float(row["jaccard"]) for row in jaccard_rows]
    fig, ax = plt.subplots(figsize=(6.8, 4.4))
    ax.bar(labels, values, color=["#0f766e", "#7c3aed", "#d97706"])
    ax.axhline(0.60, color="#b91c1c", linestyle="--", linewidth=1.0, label="60% review threshold")
    ax.set_ylim(0, 1)
    ax.set_ylabel("Jaccard overlap")
    ax.set_title("Core+probable universe versus published lists")
    ax.legend(frameon=False)
    ax.grid(axis="y", linewidth=0.35, alpha=0.3)
    fig.tight_layout()
    output.parent.mkdir(parents=True, exist_ok=True)
    save_svg(fig, output)
    plt.close(fig)


def write_notes(rows: list[dict[str, object]], jaccard_rows: list[dict[str, object]], audit_rows: list[dict[str, object]]) -> None:
    counts = Counter(str(row["surfaceome_category"]) for row in rows)
    core_probable = counts["core_surfaceome"] + counts["probable_surfaceome"]
    negative_review = [
        row
        for row in audit_rows
        if row["audit_type"] == "control"
        and row["control_role"] == "negative_intracellular_or_secreted"
        and row["surfaceome_category"] in {"core_surfaceome", "probable_surfaceome"}
    ]
    positive_excluded = [
        row
        for row in audit_rows
        if row["audit_type"] == "control" and row["control_role"] == "positive_or_benchmark" and row["surfaceome_category"] == "excluded"
    ]
    jaccard_lines = "\n".join(
        "| {published_list} | {published_n} | {intersection_n} | {union_n} | {jaccard:.4f} |".format(**row)
        for row in jaccard_rows
    )
    (DOCS_DIR / "fase4_surfaceome_universe.md").write_text(
        f"""# Fase 4 Surfaceome Universe

Access date: {dt.date.today().isoformat()}

This phase builds a conservative multi-source surfaceome universe from Cancer Surfaceome Atlas/TCSA, CSPA, SURFY, UniProt topology, UniProt reviewed lipidation/GPI-anchor evidence, UniProt GO cellular component terms, and HPA subcellular localization. The unit remains the HGNC approved protein-coding gene from Fase 3.

Core/Probable membership requires independent surface support plus an anchor/topology/localization signal: UniProt extracellular topology, confirmed UniProt lipidation `GPI-anchor`, UniProt transmembrane annotation, HPA plasma membrane localization, or SURFY surfaceome support. Curated-list or GO-only genes without this anchor are retained as ambiguous rather than used as high-confidence targets.

Confirmed UniProt lipidation `GPI-anchor` is counted as non-experimental strong anchor evidence: score +2, support source +1, `has_anchor=true`, and `has_strong=true`. Subcellular-location-only GPI annotations are flagged but not credited as confirmed direct lipid evidence.

## Category Counts

- Core surfaceome: {counts['core_surfaceome']}
- Probable surfaceome: {counts['probable_surfaceome']}
- Core + probable: {core_probable}
- Ambiguous membrane or surface context: {counts['ambiguous_membrane_or_surface_context']}
- Excluded: {counts['excluded']}

## Published-List Overlap

| Published list | Published genes | Intersection | Union | Jaccard |
|---|---:|---:|---:|---:|
{jaccard_lines}

TCSA and SURFY provide broad published-list agreement with the Core+Probable universe. CSPA overlap is lower because CSPA is an older, experimentally observed and narrower surface proteomics list; CSPA-only absences remain auditable in the false-positive/false-negative table instead of being forced into Core/Probable.

## Exit Criteria

- Negative controls in Core/Probable: {len(negative_review)}
- Positive or benchmark controls excluded: {len(positive_excluded)}
- Core + Probable below 500: {str(core_probable < 500).lower()}
- Core + Probable above 3000: {str(core_probable > 3000).lower()}

Context-dependent or non-canonical surface annotations without membrane anchor/topology support are intentionally retained as ambiguous. This prevents ER/secreted/intracellular negative controls with incidental surface annotations from entering Core/Probable while preserving them for manual review if they later appear as expression-driven top candidates.

The universe is acceptable for the next phase only if negative controls do not enter Core/Probable, positive controls are present or explained, and published-list overlap is interpretable. Genes absent from all published lists and published genes absent from Core/Probable are recorded in `surfaceome_false_positive_false_negative_audit.tsv` for top-candidate review.

## Outputs

- `data/processed/surfaceome_universe.tsv`
- `results/figures/surfaceome_source_overlap.svg`
- `results/figures/surfaceome_jaccard_with_published_lists.svg`
- `results/tables/surfaceome_confidence_summary.tsv`
- `results/tables/surfaceome_jaccard_with_published_lists.tsv`
- `results/tables/surfaceome_false_positive_false_negative_audit.tsv`
""",
        encoding="utf-8",
        newline="\n",
    )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--force-download", action="store_true")
    parser.add_argument("--timeout", type=int, default=300)
    args = parser.parse_args(argv)

    records = download_sources(force=args.force_download, timeout=args.timeout)
    update_checksum_manifests(records)
    append_provenance(records)

    universe_rows, sets, summary_rows, jaccard_rows = build_universe()
    audit_rows = build_audit_rows(universe_rows, sets)
    write_tsv(
        PROCESSED_DIR / "surfaceome_universe.tsv",
        universe_rows,
        [
            "hgnc_symbol",
            "ensembl_gene_id",
            "uniprot_accession",
            "protein_name",
            "surfaceome_confidence_score",
            "surfaceome_category",
            "surface_support_source_count",
            "surface_support_sources",
            "in_tcsa",
            "in_cspa",
            "in_surfy",
            "uniprot_extracellular_topology",
            "uniprot_gpi_anchor",
            "uniprot_gpi_evidence_class",
            "uniprot_gpi_subcellular_only",
            "uniprot_transmembrane",
            "uniprot_signal_peptide",
            "go_surface_or_plasma_membrane",
            "go_surface_terms",
            "hpa_plasma_membrane",
            "hpa_intracellular_only_flag",
            "secreted_only_flag",
            "control_role",
            "interpretation_flags",
            "tcsa_core_gesp_score",
            "tcsa_final_gesp_score",
            "cspa_category",
            "cspa_detection_cell_types",
            "surfy_label_source",
            "surfy_tm_domains",
            "hpa_main_location",
            "hpa_additional_location",
            "hpa_reliability",
            "uniprot_gpi_accession",
            "uniprot_gpi_lipidation",
        ],
    )
    write_tsv(
        RESULTS_DIR / "surfaceome_confidence_summary.tsv",
        summary_rows,
        ["summary_type", "label", "n_genes", "notes"],
    )
    write_tsv(
        RESULTS_DIR / "surfaceome_jaccard_with_published_lists.tsv",
        jaccard_rows,
        ["published_list", "own_core_probable_n", "published_n", "intersection_n", "union_n", "jaccard"],
    )
    write_tsv(
        RESULTS_DIR / "surfaceome_false_positive_false_negative_audit.tsv",
        audit_rows,
        ["audit_type", "hgnc_symbol", "surfaceome_category", "surfaceome_confidence_score", "control_role", "status", "notes"],
    )
    plot_source_overlap(summary_rows, FIGURES_DIR / "surfaceome_source_overlap.svg")
    plot_jaccard(jaccard_rows, FIGURES_DIR / "surfaceome_jaccard_with_published_lists.svg")
    write_notes(universe_rows, jaccard_rows, audit_rows)

    control_review = [row for row in audit_rows if row["audit_type"] == "control" and row["status"] == "review_required"]
    if control_review:
        print("Fase 4 failed: control audit requires review.", file=sys.stderr)
        for row in control_review:
            print(f"- {row['hgnc_symbol']}: {row['surfaceome_category']}", file=sys.stderr)
        return 1
    print("Wrote Fase 4 surfaceome universe.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
